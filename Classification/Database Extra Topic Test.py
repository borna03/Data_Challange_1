import pandas as pd
import nltk
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
import re
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import confusion_matrix, accuracy_score, classification_report
import eli5
from openpyxl import load_workbook

from sklearn.feature_extraction.text import CountVectorizer
from sklearn.pipeline import Pipeline
from sklearn.feature_extraction.text import TfidfTransformer

nltk.download('stopwords')
nltk.download('wordnet')

categories = []


def load_data():
    """
    Loads all rows from the data into a DataFrame
    :return: DataFrame with all generated AI data
    """
    global categories

    book = load_workbook('ClassificationData/Self Labeled Data2.xlsx')
    categories = book.sheetnames

    all_data_list = []
    for category in book.worksheets:
        if category.title == 'TO':
            rows = list(category.values)
            data = pd.DataFrame(rows, columns=['text', 'topic'])
            all_data_list.append(data)

            print(f"Loaded data from {category}: {data.shape[0]}")
        else:
            rows = list(category.values)
            data = pd.DataFrame(rows, columns=['text', 'topic'])
            # all_data_list.append(data)
            #
            # print(f"Loaded data from {category}: {data.shape[0]}")

    print(categories)
    self_labeled_data = pd.concat(all_data_list, ignore_index=True)

    # Clean text file
    open('Extra Topic Backup Results.txt', 'w').close()

    # Write amount of data per category in text tile
    with open('Extra Topic Backup Results.txt', 'a') as f:
        print(f'Amount of total data per category (Self labeled):', file=f)
        print('', file=f)
        print(f'Total amount of data points (Self Labeled): {self_labeled_data.shape[0]}', file=f)
        print(f'Amount of data points per category:', file=f)
        print(f'{self_labeled_data.topic.value_counts()}', file=f)
    return self_labeled_data


def preprocess(text):
    """
    Removes @'s and hyperlinks
    :param text: original text (string)
    :return: processed text
    """
    new_text = []
    for t in text.split(" "):
        t = '' if t.startswith('@') and len(t) > 1 else t
        t = '' if t.startswith('http') else t
        new_text.append(t)
    return " ".join(new_text)


def preprocess_data(df_data):
    """
    Creates a corpus for any dataframe with a 'text' column
    :param df_data: any dataframe with a 'text' column
    :return: corpus
    """
    # Remove NaN values
    df_data['text'].fillna('', inplace=True)

    # Preprocessing for Vectorization
    lemmatizer = WordNetLemmatizer()
    corpus = []

    for index, row in df_data.iterrows():
        text = row['text']
        # Remove @'s and hyperlinks
        text = preprocess(text)
        # Lowercase, lemmatizing, removing stopwords
        review = re.sub('[^A-Za-z]', ' ', text)
        review = review.lower()
        review = review.split()  # get list of words
        review = [lemmatizer.lemmatize(word) for word in review if not word in set(stopwords.words('english'))]
        review = ' '.join(review)
        corpus.append(review)

    return corpus


def train_classifier(corpus, data):
    """
    Initializes TF-IDF vectorizer, trains Logistic Regression model, extracts part of test data.
    :param corpus: corpus for data
    :param data: data (in DataFrame)
    :return: TF-IDF Vectorizer; LogisticRegression Model; Known topics for testing data; The things we want to predict
    """
    # Split the test data into a smaller portion
    X_train, X_test, y_train, y_test = train_test_split(corpus, data['topic'], test_size=0.20)

    # Initialize the vectorizer and classifier
    vec = CountVectorizer(max_features=30000)
    clf = LogisticRegression(n_jobs=1, C=1e5)

    # Compound classifier (using Pipeline)
    lr_classifier = Pipeline([('vect', vec),
                              ('tfidf', TfidfTransformer()),
                              ('clf', clf),
                              ])
    lr_classifier.fit(X_train, y_train)
    return clf, vec, lr_classifier, y_test, X_test


def classify_tweets(clf, vec, lr_classifier, y_test, X_test, data):
    """
    Predicts topics, using the Logistic Regression model, and gives some added metrics and other insights.
    :param clf: used classification model
    :param vec: used vectorizer
    :param lr_classifier: compound logistic regression classifier (pipeline class)
    :param y_test: known topics for testing data
    :param X_test: data with which we want to predict a topic
    :param data: data (in DataFrame)
    :return: DataFrame with text, true topic and predicted topic; accuracy on testing data
    """
    # Predict topic
    lr_y_pred = lr_classifier.predict(X_test)

    # Put sentence, actual label and predicted label into a DataFrame
    ind_list = y_test.index.tolist()
    df_topic_pred = data.iloc[ind_list]
    df_topic_pred.insert(2, 'predicted topic', lr_y_pred)

    # Confusion matrix
    confusion_mat = confusion_matrix(y_test, lr_y_pred)
    print(confusion_mat)

    # Accuracy test & Other evaluation metrics
    accuracy = accuracy_score(y_test, lr_y_pred)
    print(f'Accuracy of Logistic Regression on TfIdf Vectorizer data {accuracy * 100}%')

    metrics = classification_report(y_test, lr_y_pred)
    print(metrics)

    # Gets 20 most negative and 20 most positive weights for a given category
    get_weights = eli5.explain_weights_df(estimator=clf, top=(20, 20),
                                          feature_names=list(vec.get_feature_names_out())).to_string()

    # Write confusion matrix, accuracy score and DataFrame into a text tile
    with open('Extra Topic Backup Results.txt', 'a', encoding='utf-8') as f:
        print('', file=f)
        print("Confusion Matrix:", file=f)
        print(f'{confusion_mat}', file=f)
        print('', file=f)
        print(f'Accuracy of Logistic Regression on TfIdf Vectorizer data: {accuracy * 100}%', file=f)
        print('', file=f)
        print(f'Metrics: \n{metrics}', file=f)
        print('Precision = predicted topic correct / amount of times a topic was predicted', file=f)
        print('Recall = predicted topic correct / amount of true (labeled) topic classifications', file=f)
        print('Support = amount of true (labeled) topic classifications', file=f)
        print(
            "So good precision, but bad recall ==> model doesn't recognize a topic often, but when it does, it does so correctly",
            file=f)
        print(
            "So good recall, but bad precision ==> model predicts this topic too much, therefore also gets it right when it's the actual topic",
            file=f)
        print('', file=f)
        print(f'{df_topic_pred.to_string()}', file=f)
        print('', file=f)
        print(f'{get_weights}', file=f)

    return df_topic_pred, accuracy


def run():
    """
    Run all functions for TF-IDF Vectorization + Logistic Regression tests. Model trained on AI data,
    and tested on human labeled data.
    """
    # Load data into DataFrames
    data = load_data()
    print(data)

    # Preprocess data (into corpus)
    corpus = preprocess_data(data)

    # Train TF-IDF and Logistic Regression on the AI data
    clf, vec, lr_classifier, y_test, X_test = train_classifier(corpus, data)

    # Predict topics
    df_topic_pred, accuracy = classify_tweets(clf, vec, lr_classifier, y_test, X_test, data)
    print(df_topic_pred.to_string())

    # Show Vectorization weights
    print(eli5.explain_weights_df(estimator=clf, top=(20, 20),
                                  feature_names=list(vec.get_feature_names_out())).to_string())


def accuracy_run(runs):
    """
    Gets average accuracy over multiples runs.
    Per run, it will call all functions for TF-IDF Vectorization + Logistic Regression tests.
    Model trained on AI data,
    and tested on human labeled data.
    :param runs: amount of runs
    """
    total_acc = 0
    count = 0
    for i in range(0, runs):
        print(f'Current iteration: {count}')
        count += 1

        # Load data into DataFrames
        data = load_data()

        # Preprocess data (into corpus)
        corpus = preprocess_data(data)

        # Train TF-IDF and Logistic Regression on the AI data
        clf, vec, lr_classifier, y_test, X_test = train_classifier(corpus, data)

        # Predict topics
        df_topic_pred, accuracy = classify_tweets(clf, vec, lr_classifier, y_test, X_test, data)
        print(df_topic_pred)

        print(f'Accuracy: {accuracy}')
        total_acc += accuracy
        print('\n')

    print('\n')
    print(f'Accumulated accuracy: {total_acc}')
    print(f'Total iterations: {count}')
    print(f'Average accuracy: {total_acc / count * 100}')

    with open('Extra Topic Accuracy Backup Results.txt', 'a', encoding='utf-8') as f:
        print(
            f'Trained on self-labeled data, tested on pre-labeled data (Current Classifier Model | Count Vectorizer | 244 Data Points)',
            file=f)
        print(f'Accumulated accuracy: {total_acc}', file=f)
        print(f'Total iterations: {count}', file=f)
        print(f'Average accuracy: {total_acc / count * 100}', file=f)


run()
# accuracy_run(20)
