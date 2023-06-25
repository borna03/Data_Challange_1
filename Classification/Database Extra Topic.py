from data_load import *
from tqdm import tqdm
import time

import pandas as pd
import nltk
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
import re
from sklearn.linear_model import LogisticRegression
from openpyxl import load_workbook
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.pipeline import Pipeline
from sklearn.feature_extraction.text import TfidfTransformer

nltk.download('stopwords')
nltk.download('wordnet')

categories = []


def load_training_data():
    """
    Loads all rows from the data into a DataFrame
    :return: DataFrame with all generated AI data
    """
    global categories

    book = load_workbook('ClassificationData/Self Labeled Data2.xlsx')
    categories = book.sheetnames

    all_data_list = []
    for category in book.worksheets:
        if category.title == 'TO':
            rows = list(category.values)
            data = pd.DataFrame(rows, columns=['text', 'topic'])
            all_data_list.append(data)

            print(f"Loaded data from {category}: {data.shape[0]}")
        else:
            None

    print(categories)
    self_labeled_data = pd.concat(all_data_list, ignore_index=True)

    # Clean text file
    open('Extra Topic Backup Results.txt', 'w').close()

    # Write amount of data per category in text tile
    with open('Extra Topic Backup Results.txt', 'a') as f:
        print(f'Amount of total data per category (Self labeled):', file=f)
        print('', file=f)
        print(f'Total amount of data points (Self Labeled): {self_labeled_data.shape[0]}', file=f)
        print(f'Amount of data points per category:', file=f)
        print(f'{self_labeled_data.topic.value_counts()}', file=f)
    return self_labeled_data


def preprocess(text):
    """
    Removes @'s and hyperlinks
    :param text: original text (string)
    :return: processed text
    """
    new_text = []
    for t in text.split(" "):
        t = '' if t.startswith('@') and len(t) > 1 else t
        t = '' if t.startswith('http') else t
        new_text.append(t)
    return " ".join(new_text)


def preprocess_text(text):
    """
    Creates a corpus for any text string
    :param text: any text (string)
    :return: corpus of the input text
    """
    # Preprocessing for Vectorization
    lemmatizer = WordNetLemmatizer()
    corpus = []

    # Remove @'s and hyperlinks
    text = preprocess(text)
    # Lowercase, lemmatizing, removing stopwords
    review = re.sub('[^A-Za-z]', ' ', text)
    review = review.lower()
    review = review.split()  # get list of words
    review = [lemmatizer.lemmatize(word) for word in review if not word in set(stopwords.words('english'))]
    review = ' '.join(review)
    corpus.append(review)

    return corpus


def preprocess_df_data(df_data):
    """
    Creates a corpus for any dataframe with a 'text' column
    :param df_data: any dataframe with a 'text' column
    :return: corpus
    """
    # Remove NaN values
    df_data['text'].fillna('', inplace=True)

    # Preprocessing for Vectorization
    lemmatizer = WordNetLemmatizer()
    corpus = []

    for index, row in df_data.iterrows():
        text = row['text']
        # Remove @'s and hyperlinks
        text = preprocess(text)
        # Lowercase, lemmatizing, removing stopwords
        review = re.sub('[^A-Za-z]', ' ', text)
        review = review.lower()
        review = review.split()  # get list of words
        review = [lemmatizer.lemmatize(word) for word in review if not word in set(stopwords.words('english'))]
        review = ' '.join(review)
        corpus.append(review)

    return corpus


def train_classifier(training_corpus, training_data):
    """
    Trains the Logistic Regression model, using the self-labeled tweets
    :param training_corpus: corpus for the self-labeled data
    :param training_data: self-labeled data (in DataFrame)
    :return: LogisticRegression Model
    """
    # Initialize the vectorizer and classifier
    vec = CountVectorizer(max_features=30000)
    clf = LogisticRegression(n_jobs=1, C=1e5)

    # Compound classifier (using Pipeline)
    lr_classifier = Pipeline([('vect', vec),
                              ('tfidf', TfidfTransformer()),
                              ('clf', clf),
                              ])
    lr_classifier.fit(training_corpus, training_data['topic'])
    return lr_classifier


def classify_tweets(lr_classifier, id_str):
    """
    Predicts topics, using the Logistic Regression model, and writes this into the MongoDB database.
    :param lr_classifier: compound logistic regression classifier (pipeline class)
    :param id_str: id (in string form) of the wanted airline
    """
    total_count = 0
    error_count = 0
    start_time = time.time()

    # Find all tweets for which the airline is mentioned (@) or is replied to:
    query = {'$and': [
        {'topic_specific': None},
        {'$or': [{'user.id_str': id_str}, {'in_reply_to_user_id_str': id_str}, {'entities.user_mentions': {'$elemMatch': {'id_str': id_str}}}]}
    ]}

    print(f'Expected amount of iterations/tweets: {collection.count_documents(query)}')

    all_tweets = collection.find(query)

    print('Starting analysis:')
    # Iterate over all tweets, with live counter
    for tweet in tqdm(all_tweets):
        try:
            total_count += 1
            text = tweet['text']
            processed_text = preprocess_text(text)
            lr_topic_pred = lr_classifier.predict(processed_text)

            # Put topic prediction as value of 'topic_specific' key
            newvalues_topic = {'$set': {'topic_specific': lr_topic_pred[0]}}
            collection.update_one(tweet, newvalues_topic)
        except Exception as e:
            print(e)
            error_count += 1

    print(f'Error count: {error_count}')
    print('\n')
    print(f'Total evaluated tweets: {total_count}')
    print('\n')

    end_time = time.time()
    elapsed_time = end_time - start_time
    print("Time: ", elapsed_time)
    print("Time/Item: ", elapsed_time / total_count)


def run():
    """
    Run all functions for TF-IDF Vectorization + Logistic Regression tests. Model trained on AI data,
    and tested on human labeled data.
    """
    # Load data into DataFrames
    training_data = load_training_data()
    print(training_data)

    # Process training data into a corpus
    training_corpus = preprocess_df_data(training_data)
    print(training_corpus)

    # Train Logistic Regression model
    lr_classifier = train_classifier(training_corpus, training_data)

    print(lr_classifier)

    # Predict topics
    classify_tweets(lr_classifier, '18332190')

run()
