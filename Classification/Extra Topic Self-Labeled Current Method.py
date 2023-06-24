import pandas as pd
import nltk
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
import re
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import confusion_matrix, accuracy_score, classification_report
import eli5
from openpyxl import load_workbook

from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer
from sklearn.pipeline import Pipeline
from sklearn.feature_extraction.text import TfidfTransformer

nltk.download('stopwords')
nltk.download('wordnet')

categories = []


def load_training_data():
    """
    Loads all rows from the AI generated data into a DataFrame
    :return: DataFrame with all generated AI data
    """
    global categories

    book = load_workbook('ClassificationData/Self Labeled Data.xlsx')
    categories = book.sheetnames

    all_data_list = []
    for category in book.worksheets:
        if category.title == 'TO':
            rows = list(category.values)
            data = pd.DataFrame(rows, columns=['text', 'topic'])
            all_data_list.append(data)

            print(f"Loaded data from {category}: {data.shape[0]}")
        else:
            rows = list(category.values)
            data = pd.DataFrame(rows, columns=['text', 'topic'])
            # all_data_list.append(data)
            #
            # print(f"Loaded data from {category}: {data.shape[0]}")

    print(categories)
    self_labeled_data = pd.concat(all_data_list, ignore_index=True)

    self_labeled_data = self_labeled_data.loc[(self_labeled_data['topic'] == 'Customer Service') |
                                              (self_labeled_data['topic'] == 'Delays & Cancellations') |
                                              (self_labeled_data['topic'] == 'Baggage') |
                                              (self_labeled_data['topic'] == 'Flight Experience') |
                                              (self_labeled_data['topic'] == "Can't Tell") |
                                              (self_labeled_data['topic'] == 'Booking & Reservation') |
                                              (self_labeled_data['topic'] == 'Long Lines')]

    # Clean text file
    open('Extra Topic Backup Results.txt', 'w').close()

    # Write amount of data per category in text tile
    with open('Extra Topic Backup Results.txt', 'a') as f:
        print(f'Amount of total data per category (Self labeled):', file=f)
        print('', file=f)
        print(f'Total amount of data points (Self Labeled): {self_labeled_data.shape[0]}', file=f)
        print(f'Amount of data points per category:', file=f)
        print(f'{self_labeled_data.topic.value_counts()}', file=f)
    return self_labeled_data


def load_test_data():
    """
    Loads the useful rows from the human labeled dataset into a DataFrame
    :return: DataFrame with necessary rows from the human labeled dataset
    """
    # Load the data in the labeled csv file as a pandas DataFrame
    csv_file = f'ClassificationData/HumanLabeledData.csv'
    data = pd.read_csv(csv_file, delimiter=',', header=0,
                       usecols=['airline_sentiment', 'airline_sentiment_confidence', 'negativereason',
                                'negativereason_confidence', 'text'])
    data.rename(columns={'negativereason': 'topic', 'negativereason_confidence': 'topic_confidence'}, inplace=True)

    # Masks (loc) to get the needed rows
    data = data.loc[data['airline_sentiment'] == 'negative'].loc[data['airline_sentiment_confidence'] > 0.75].loc[
        data['topic_confidence'] > 0.75]

    # Reset index to start with 0, drop unneeded columns
    human_labeled_data = data.reset_index().drop(
        columns=['index', 'airline_sentiment', 'airline_sentiment_confidence', 'topic_confidence'])

    # Combining and re-naming topics
    human_labeled_data['topic'].mask(
        ((human_labeled_data['topic'] == 'Damaged Luggage') | (human_labeled_data['topic'] == 'Lost Luggage')),
        'Baggage', inplace=True)
    human_labeled_data['topic'].mask(((human_labeled_data['topic'] == 'Bad Flight') | (
            human_labeled_data['topic'] == 'Flight Attendant Complaints')), 'Flight Experience', inplace=True)
    human_labeled_data['topic'].mask(
        ((human_labeled_data['topic'] == 'Late Flight') | (human_labeled_data['topic'] == 'Cancelled Flight')),
        'Delays & Cancellations', inplace=True)
    human_labeled_data['topic'].mask((human_labeled_data['topic'] == 'Customer Service Issue'), 'Customer Service',
                                     inplace=True)
    human_labeled_data['topic'].mask((human_labeled_data['topic'] == 'longlines'), 'Long Lines', inplace=True)
    human_labeled_data['topic'].mask((human_labeled_data['topic'] == 'Flight Booking Problems'),
                                     'Booking & Reservation', inplace=True)

    # Write amount of data points in file
    with open('Extra Topic Backup Results.txt', 'a') as f:
        print('', file=f)
        print(f'Total amount of data points (Human Labeled): {human_labeled_data.shape[0]}', file=f)
        print(f'Amount of data points per category:', file=f)
        print(f'{human_labeled_data.topic.value_counts(normalize=True)}', file=f)

    return human_labeled_data


def preprocess(text):
    """
    Removes @'s and hyperlinks
    :param text: original text (string)
    :return: processed text
    """
    new_text = []
    for t in text.split(" "):
        t = '' if t.startswith('@') and len(t) > 1 else t
        t = '' if t.startswith('http') else t
        new_text.append(t)
    return " ".join(new_text)


def preprocess_data(df_data):
    """
    Creates a corpus for any dataframe with a 'text' column
    :param df_data: any dataframe with a 'text' column
    :return: corpus
    """
    # Remove NaN values
    df_data['text'].fillna('', inplace=True)

    # Preprocessing for Vectorization
    lemmatizer = WordNetLemmatizer()
    corpus = []

    for index, row in df_data.iterrows():
        text = row['text']
        # Remove @'s and hyperlinks
        text = preprocess(text)
        # Lowercase, lemmatizing, removing stopwords
        review = re.sub('[^A-Za-z]', ' ', text)
        review = review.lower()
        review = review.split()  # get list of words
        review = [lemmatizer.lemmatize(word) for word in review if not word in set(stopwords.words('english'))]
        review = ' '.join(review)
        corpus.append(review)

    return corpus


def train_classifier(corpus_train, corpus_test, train_data, test_data):
    """
    Initializes TF-IDF vectorizer, trains Logistic Regression model, extracts part of test data.
    :param corpus_train: corpus for training data
    :param corpus_test: corpus for test data
    :param train_data: training data (in DataFrame)
    :param test_data: test data (in DataFrame)
    :return: TF-IDF Vectorizer; LogisticRegression Model; Known topics for testing data; The things we want to predict
    """
    # Split the test data into a smaller portion
    X_train, X_test, y_train, y_test = train_test_split(corpus_test, test_data['topic'], test_size=0.05)

    # Initialize the vectorizer and classifier
    vec = CountVectorizer(max_features=30000)
    clf = LogisticRegression(n_jobs=1, C=1e5)

    # Compound classifier (using Pipeline)
    lr_classifier = Pipeline([('vect', vec),
                              ('tfidf', TfidfTransformer()),
                              ('clf', clf),
                              ])
    lr_classifier.fit(corpus_train, train_data['topic'])
    return clf, vec, lr_classifier, y_test, X_test


def classify_tweets(clf, vec, lr_classifier, y_test, X_test, test_data):
    """
    Predicts topics, using the Logistic Regression model, and gives some added metrics and other insights.
    :param clf: used classification model
    :param vec: used vectorizer
    :param lr_classifier: compound logistic regression classifier (pipeline class)
    :param y_test: known topics for testing data
    :param X_test: data with which we want to predict a topic
    :param test_data: test data (in DataFrame)
    :return: DataFrame with text, true topic and predicted topic; accuracy on testing data
    """
    # Predict topic
    lr_y_pred = lr_classifier.predict(X_test)

    # Put sentence, actual label and predicted label into a DataFrame
    ind_list = y_test.index.tolist()
    df_topic_pred = test_data.iloc[ind_list]
    df_topic_pred.insert(2, 'predicted topic', lr_y_pred)

    # Confusion matrix
    confusion_mat = confusion_matrix(y_test, lr_y_pred)
    print(confusion_mat)

    # Accuracy test & Other evaluation metrics
    accuracy = accuracy_score(y_test, lr_y_pred)
    print(f'Accuracy of Logistic Regression on TfIdf Vectorizer data {accuracy * 100}%')

    metrics = classification_report(y_test, lr_y_pred)
    print(metrics)

    # Gets 20 most negative and 20 most positive weights for a given category
    get_weights = eli5.explain_weights_df(estimator=clf, top=(20, 20),
                                          feature_names=list(vec.get_feature_names_out())).to_string()

    # Write confusion matrix, accuracy score and DataFrame into a text tile
    with open('Extra Topic Backup Results.txt', 'a', encoding='utf-8') as f:
        print('', file=f)
        print("Confusion Matrix:", file=f)
        print(f'{confusion_mat}', file=f)
        print('', file=f)
        print(f'Accuracy of Logistic Regression on TfIdf Vectorizer data: {accuracy * 100}%', file=f)
        print('', file=f)
        print(f'Metrics: \n{metrics}', file=f)
        print('Precision = predicted topic correct / amount of times a topic was predicted', file=f)
        print('Recall = predicted topic correct / amount of true (labeled) topic classifications', file=f)
        print('Support = amount of true (labeled) topic classifications', file=f)
        print(
            "So good precision, but bad recall ==> model doesn't recognize a topic often, but when it does, it does so correctly",
            file=f)
        print(
            "So good recall, but bad precision ==> model predicts this topic too much, therefore also gets it right when it's the actual topic",
            file=f)
        print('', file=f)
        print(f'{df_topic_pred.to_string()}', file=f)
        print('', file=f)
        print(f'{get_weights}', file=f)

    return df_topic_pred, accuracy


def run():
    """
    Run all functions for TF-IDF Vectorization + Logistic Regression tests. Model trained on AI data,
    and tested on human labeled data.
    """
    # Load data into DataFrames
    train_data = load_training_data()
    test_data = load_test_data()
    print(train_data)
    print(test_data)

    # Preprocess data (into corpus)
    corpus_train = preprocess_data(train_data)
    corpus_test = preprocess_data(test_data)

    # Train TF-IDF and Logistic Regression on the AI data
    clf, vec, lr_classifier, y_test, X_test = train_classifier(corpus_train, corpus_test, train_data,
                                                               test_data)

    # Predict topics
    df_topic_pred, accuracy = classify_tweets(clf, vec, lr_classifier, y_test, X_test, test_data)
    print(df_topic_pred.to_string())

    # Show Vectorization weights
    print(eli5.explain_weights_df(estimator=clf, top=(20, 20),
                                  feature_names=list(vec.get_feature_names_out())).to_string())


def accuracy_run(runs):
    """
    Gets average accuracy over multiples runs.
    Per run, it will call all functions for TF-IDF Vectorization + Logistic Regression tests.
    Model trained on AI data,
    and tested on human labeled data.
    :param runs: amount of runs
    """
    total_acc = 0
    count = 0
    for i in range(0, runs):
        count += 1

        # Load data into DataFrames
        train_data = load_training_data()
        test_data = load_test_data()

        # Preprocess data (into corpus)
        corpus_train = preprocess_data(train_data)
        corpus_test = preprocess_data(test_data)

        # Train TF-IDF and Logistic Regression on the AI data
        clf, vec, lr_classifier, y_test, X_test = train_classifier(corpus_train, corpus_test, train_data,
                                                                   test_data)

        # Predict topics
        df_topic_pred, accuracy = classify_tweets(clf, vec, lr_classifier, y_test, X_test, test_data)
        print(df_topic_pred)

        print(f'Accuracy: {accuracy}')
        total_acc += accuracy

    print('\n')
    print(f'Accumulated accuracy: {total_acc}')
    print(f'Total iterations: {count}')
    print(f'Average accuracy: {total_acc / count * 100}')

    with open('Extra Topic Accuracy Backup Results.txt', 'a', encoding='utf-8') as f:
        print(f'Trained on self-labeled data, trained on pre-labeled data', file=f)
        print(f'Accumulated accuracy: {total_acc}', file=f)
        print(f'Total iterations: {count}', file=f)
        print(f'Average accuracy: {total_acc / count * 100}', file=f)


# run()
# accuracy_run(5)
