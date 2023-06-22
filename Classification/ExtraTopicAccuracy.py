import pandas as pd
import nltk
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
import re
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import confusion_matrix, accuracy_score, classification_report
import eli5
from openpyxl import Workbook, load_workbook

nltk.download('stopwords')
nltk.download('wordnet')

categories = []
def load_classification_data():
    global categories
    # Clean text file
    open('Extra Topic Backup Results.txt', 'w').close()

    # Write something in text file
    with open('Extra Topic Backup Results.txt', 'a') as f:
        print(f'Amount of total data per category:', file=f)

    book = load_workbook('ClassificationData/AllData.xlsx')
    categories = book.sheetnames

    all_data_list = []
    for category in book.worksheets:
        rows = list(category.values)
        data = pd.DataFrame(rows, columns=['text', 'topic'])
        all_data_list.append(data)

        print(f"Loaded data from {category}: {data.shape[0]}")

        # Write total amount of data point per category in file
        with open('Extra Topic Backup Results.txt', 'a') as f:
            print(f'{category}: {data.shape[0]}', file=f)
    print(categories)
    combined_data = pd.concat(all_data_list, ignore_index=True)
    return combined_data

def load_classification_data2():
    # Load the data in the labeled csv file as a pandas DataFrame
    csv_file = f'ClassificationData/HumanLabeledData.csv'
    data = pd.read_csv(csv_file, delimiter=',', header=0,
                       usecols=['airline_sentiment', 'airline_sentiment_confidence', 'negativereason',
                                'negativereason_confidence', 'text'])
    data.rename(columns={'negativereason': 'topic', 'negativereason_confidence': 'topic_confidence'}, inplace=True)

    # Masks (loc) to get the needed rows
    data = data.loc[data['airline_sentiment'] == 'negative'].loc[data['airline_sentiment_confidence'] > 0.75].loc[
        data['topic_confidence'] > 0.75]

    # Reset index to start with 0, drop unneeded columns
    cleaned_data = data.reset_index().drop(
        columns=['index', 'airline_sentiment', 'airline_sentiment_confidence', 'topic_confidence'])

    # # Shuffle the data and assign arbitrary seed
    # cleaned_data = cleaned_data.sample(frac=1, random_state=random.randint(0, 1000))

    # Clean text file
    open('Extra Topic Backup Results.txt', 'w').close()

    # Write amount of data points in file
    with open('Extra Topic Backup Results.txt', 'a') as f:
        print(f'Total amount of data points: {cleaned_data.shape[0]}', file=f)
        print('', file=f)
        print(f'Amount of data points per category:', file=f)
        print(f'{cleaned_data.topic.value_counts(normalize=True)}', file=f)

    return cleaned_data


def preprocess(text):
    new_text = []
    for t in text.split(" "):
        t = '' if t.startswith('@') and len(t) > 1 else t
        t = '' if t.startswith('http') else t
        new_text.append(t)
    return " ".join(new_text)


def preprocess_data(combined_data):
    # Remove NaN values
    combined_data['text'].fillna('', inplace=True)

    # Preprocessing for Vectorization
    lemmatizer = WordNetLemmatizer()
    corpus = []

    for index, row in combined_data.iterrows():
        text = row['text']
        # Remove @'s and hyperlinks
        text = preprocess(text)
        # Lowercase, lemmatizing, removing stopwords
        review = re.sub('[^A-Za-z]', ' ', text)
        review = review.lower()
        review = review.split()  # get list of words
        review = [lemmatizer.lemmatize(word) for word in review if not word in set(stopwords.words('english'))]
        review = ' '.join(review)
        corpus.append(review)

    return corpus


def train_classifier(corpus, combined_data):
    # Initialize TfidfVectorizer
    tfidf = TfidfVectorizer(ngram_range=(2, 2), max_features=100000)
    X_tf = tfidf.fit_transform(corpus).toarray()

    # Split data into training and test data
    X_train, X_test, y_train, y_test = train_test_split(X_tf, combined_data['topic'], test_size=0.25)

    # Initialize Logistic Regression Model
    lr_classifier = LogisticRegression(C=5e1, solver='lbfgs', multi_class='multinomial', n_jobs=4)
    lr_classifier.fit(X_train, y_train)

    return tfidf, lr_classifier, y_test, X_test


def classify_tweets(lr_classifier, tfidf, y_test, X_test, combined_data, data_type):
    # Predict topic
    lr_y_pred = lr_classifier.predict(X_test)

    # Put sentence, actual label and predicted label into a DataFrame
    ind_list = y_test.index.tolist()
    df_topic_pred = combined_data.iloc[ind_list]
    df_topic_pred.insert(2, 'predicted topic', lr_y_pred)

    # Confusion matrix
    confusion_mat = confusion_matrix(y_test, lr_y_pred)
    print(confusion_mat)

    # Accuracy test & Other evaluation metrics
    accuracy = accuracy_score(y_test, lr_y_pred)
    print(f'Accuracy of Logistic Regression on TfIdf Vectorizer data {accuracy * 100}%')
    if data_type == 1:
        topics = categories
    elif data_type == 2:
        topics = ['Bad Flight', "Can't Tell", 'Cancelled Flight', 'Customer Service Issue', 'Damaged Luggage',
                  'Flight Attendant Complaints', 'Flight Booking Problems', 'Late Flight', 'Lost Luggage', 'longlines']

    metrics = classification_report(y_test, lr_y_pred, target_names=topics)
    print(metrics)

    # Gets 20 most negative and 20 most positive weights for a given category
    get_weights = eli5.explain_weights_df(estimator=lr_classifier, top=(20, 20),
                                          feature_names=list(tfidf.get_feature_names_out())).to_string()

    # Write confusion matrix, accuracy score and DataFrame into a text tile
    with open('Extra Topic Backup Results.txt', 'a') as f:
        print('', file=f)
        print(f'{confusion_mat}', file=f)
        print('', file=f)
        print(f'Accuracy of Logistic Regression on TfIdf Vectorizer data: {accuracy * 100}%', file=f)
        print('', file=f)
        print(f'Metrics: \n{metrics}', file=f)

        try:
            print('', file=f)
            print(f'{df_topic_pred.to_string()}', file=f)
            print('', file=f)
            print(f'{get_weights}', file=f)
        except UnicodeEncodeError:
            print('UnicodeEncodeError')

    return df_topic_pred, accuracy, lr_y_pred


def run(data_type):
    """
    Run all functions for TF-IDF Vectorization + Logistic Regression tests, either for AI generated or human labeled data.
    :param data_type: for AI data, input 1, for human labeled data, input 2
    :return: invalid input message, if applicable
    """
    # Load data into DataFrame
    if data_type == 1:
        combined_data = load_classification_data()
    elif data_type == 2:
        combined_data = load_classification_data2()
    else:
        return 'Invalid input'

    print(combined_data)

    # Preprocess data (into corpus)
    corpus = preprocess_data(combined_data)

    # Train TF-IDF and Logistic Regression
    tfidf, lr_classifier, y_test, X_test = train_classifier(corpus, combined_data)

    # Predict topics
    df_topic_pred, accuracy, lr_y_pred = classify_tweets(lr_classifier, tfidf, y_test, X_test, combined_data, data_type)
    print(df_topic_pred.to_string())

    # Show Vectorization weights
    print(eli5.explain_weights_df(estimator=lr_classifier, top=(20, 20),
                                  feature_names=list(tfidf.get_feature_names_out())).to_string())


# run(2)

def accuracy_run(runs, data_type):
    """
    Get average accuracy over multiple runs, either for AI generated or human labeled data.
    :param runs: amount of runs to get an average accuracy from
    :param data_type: for AI data, input 1, for human labeled data, input 2
    """

    total_acc = 0
    count = 0
    for i in range(0, runs):
        count += 1
        # Load data into DataFrame
        if data_type == 1:
            combined_data = load_classification_data()
        elif data_type == 2:
            combined_data = load_classification_data2()
        else:
            return 'Invalid input'

        print(combined_data)

        # Preprocess data (into corpus)
        corpus = preprocess_data(combined_data)

        # Train TF-IDF and Logistic Regression
        tfidf, lr_classifier, y_test, X_test = train_classifier(corpus, combined_data)

        # Predict topics
        df_topic_pred, accuracy, lr_y_pred = classify_tweets(lr_classifier, tfidf, y_test, X_test, combined_data,
                                                             data_type)

        print(f'Accuracy: {accuracy}')
        total_acc += accuracy

    print('\n')
    print(f'Accumulated accuracy: {total_acc}')
    print(f'Total iterations: {count}')
    print(f'Average accuracy: {total_acc / count}')

data = load_classification_data()
print(data)