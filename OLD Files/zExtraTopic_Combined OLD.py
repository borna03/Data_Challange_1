import pandas as pd
import nltk
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
import re
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import confusion_matrix, accuracy_score, classification_report
import eli5
from openpyxl import Workbook, load_workbook

nltk.download('stopwords')
nltk.download('wordnet')

def load_classification_data():
    # Load AI Data
    book = load_workbook('../Classification/ClassificationData/AllData.xlsx')
    categories = book.sheetnames

    all_data_list = []
    for category in book.worksheets:
        rows = list(category.values)
        data = pd.DataFrame(rows, columns=['text', 'topic'])
        all_data_list.append(data)

        print(f"Loaded data from {category}: {data.shape[0]}")

        # Write total amount of data point per category in file
        with open('../Classification/Extra Topic Backup Results.txt', 'a') as f:
            print(f'{category}: {data.shape[0]}', file=f)
    AI_data_full = pd.concat(all_data_list, ignore_index=True)

    # Load Human Labeled Data
    # Load the data in the labeled csv file as a pandas DataFrame
    csv_file = f'../Classification/ClassificationData/HumanLabeledData.csv'
    data_labeled = pd.read_csv(csv_file, delimiter=',', header=0,
                       usecols=['airline_sentiment', 'airline_sentiment_confidence', 'negativereason',
                                'negativereason_confidence', 'text'])
    data_labeled.rename(columns={'negativereason': 'topic', 'negativereason_confidence': 'topic_confidence'}, inplace=True)

    # Masks (loc) to get the needed rows; masks to group topics together (lessen category imbalance)
    data_labeled = data_labeled.loc[data_labeled['airline_sentiment'] == 'negative'].loc[data_labeled['airline_sentiment_confidence'] > 0.75].loc[
        data_labeled['topic_confidence'] > 0.75]


    # Combine Data
    combined_data = pd.concat([AI_data_full, data_labeled]).reset_index().drop(
        columns=['index', 'airline_sentiment', 'airline_sentiment_confidence', 'topic_confidence'])

    # Combine topics (lessen category imbalance)
    combined_data['topic'].mask(((combined_data['topic'] == 'Damaged Luggage') | (combined_data['topic'] == 'Lost Luggage') | (combined_data['topic'] == 'Baggage and Security')), 'Luggage Problems', inplace=True)
    combined_data['topic'].mask(((combined_data['topic'] == 'Bad Flight') | (combined_data['topic'] == 'Flight Attendant Complaints') | (combined_data['topic'] == 'Flight Experience')), 'On-flight Problems', inplace=True)
    combined_data['topic'].mask(((combined_data['topic'] == 'Late Flight') | (combined_data['topic'] == 'Cancelled Flight')), 'Delayed/Cancelled Flight', inplace=True)
    combined_data['topic'].mask(((combined_data['topic'] == 'Customer Service') | (combined_data['topic'] == 'Customer Service Issue')), 'Customer Service Remarks', inplace=True)

    # Clean text file
    open('../Classification/Extra Topic Backup Results.txt', 'w').close()

    # Write amount of data points in file
    with open('../Classification/Extra Topic Backup Results.txt', 'a') as f:
        print(f'Total amount of data points: {combined_data.shape[0]}', file=f)
        print('', file=f)
        print(f'Amount of data points per category:', file=f)
        print(f'{combined_data.topic.value_counts(normalize=True)}', file=f)

    return combined_data

print(load_classification_data())

def preprocess(text):
    new_text = []
    for t in text.split(" "):
        t = '' if t.startswith('@') and len(t) > 1 else t
        t = '' if t.startswith('http') else t
        new_text.append(t)
    return " ".join(new_text)


def preprocess_data(combined_data):
    # Remove NaN values
    combined_data['text'].fillna('', inplace=True)

    # Preprocessing for Vectorization
    lemmatizer = WordNetLemmatizer()
    corpus = []

    for index, row in combined_data.iterrows():
        text = row['text']
        # Remove @'s and hyperlinks
        text = preprocess(text)
        # Lowercase, lemmatizing, removing stopwords
        review = re.sub('[^A-Za-z]', ' ', text)
        review = review.lower()
        review = review.split()  # get list of words
        review = [lemmatizer.lemmatize(word) for word in review if not word in set(stopwords.words('english'))]
        review = ' '.join(review)
        corpus.append(review)

    return corpus


def train_classifier(corpus, combined_data):
    # Initialize TfidfVectorizer
    tfidf = TfidfVectorizer(ngram_range=(2, 2), max_features=100000)
    X_tf = tfidf.fit_transform(corpus).toarray()

    # Split data into training and test data
    X_train, X_test, y_train, y_test = train_test_split(X_tf, combined_data['topic'], test_size=0.25)

    # Initialize Logistic Regression Model
    lr_classifier = LogisticRegression(C=5e1, solver='lbfgs', multi_class='multinomial', n_jobs=4)
    lr_classifier.fit(X_train, y_train)

    return tfidf, lr_classifier, y_test, X_test


def classify_tweets(lr_classifier, tfidf, y_test, X_test, combined_data):
    # Predict topic
    lr_y_pred = lr_classifier.predict(X_test)

    # Put sentence, actual label and predicted label into a DataFrame
    ind_list = y_test.index.tolist()
    df_topic_pred = combined_data.iloc[ind_list]
    df_topic_pred.insert(2, 'predicted topic', lr_y_pred)

    # Confusion matrix
    confusion_mat = confusion_matrix(y_test, lr_y_pred)
    print(confusion_mat)

    # Accuracy test & Other evaluation metrics
    accuracy = accuracy_score(y_test, lr_y_pred)
    print(f'Accuracy of Logistic Regression on TfIdf Vectorizer data {accuracy * 100}%')
    topics = ['On-flight problems', 'Customer Service Issue', 'Delayed/Cancelled Flight', 'Flight Booking Problems', 'Luggage Problems',
                  "Can't tell", 'Long Lines']

    metrics = classification_report(y_test, lr_y_pred, target_names=topics)
    print(metrics)

    # Gets 20 most negative and 20 most positive weights for a given category
    get_weights = eli5.explain_weights_df(estimator=lr_classifier, top=(20, 20),
                                          feature_names=list(tfidf.get_feature_names_out())).to_string()

    # Write confusion matrix, accuracy score and DataFrame into a text tile
    with open('../Classification/Extra Topic Backup Results.txt', 'a') as f:
        print('', file=f)
        print(f'{confusion_mat}', file=f)
        print('', file=f)
        print(f'Accuracy of Logistic Regression on TfIdf Vectorizer data: {accuracy * 100}%', file=f)
        print('', file=f)
        print(f'Metrics: \n{metrics}', file=f)

        try:
            print('', file=f)
            print(f'{df_topic_pred.to_string()}', file=f)
            print('', file=f)
            print(f'{get_weights}', file=f)
        except UnicodeEncodeError:
            print('UnicodeEncodeError')

    return df_topic_pred, accuracy, lr_y_pred


def run():
    """
    Run all functions for TF-IDF Vectorization + Logistic Regression tests, either for AI generated or human labeled data.
    :param data_type: for AI data, input 1, for human labeled data, input 2
    :return: invalid input message, if applicable
    """

    combined_data = load_classification_data()
    print(combined_data)

    # Preprocess data (into corpus)
    corpus = preprocess_data(combined_data)

    # Train TF-IDF and Logistic Regression
    tfidf, lr_classifier, y_test, X_test = train_classifier(corpus, combined_data)

    # Predict topics
    df_topic_pred, accuracy, lr_y_pred = classify_tweets(lr_classifier, tfidf, y_test, X_test, combined_data)
    print(df_topic_pred.to_string())

    # Show Vectorization weights
    print(eli5.explain_weights_df(estimator=lr_classifier, top=(20, 20),
                                  feature_names=list(tfidf.get_feature_names_out())).to_string())


run()

def accuracy_run(runs):
    """
    Get average accuracy over multiple runs, either for AI generated or human labeled data.
    :param runs: amount of runs to get an average accuracy from
    :param data_type: for AI data, input 1, for human labeled data, input 2
    """

    total_acc = 0
    count = 0
    for i in range(0, runs):
        count += 1
        # Load data into DataFrame
        combined_data = load_classification_data()
        print(combined_data)

        # Preprocess data (into corpus)
        corpus = preprocess_data(combined_data)

        # Train TF-IDF and Logistic Regression
        tfidf, lr_classifier, y_test, X_test = train_classifier(corpus, combined_data)

        # Predict topics
        df_topic_pred, accuracy, lr_y_pred = classify_tweets(lr_classifier, tfidf, y_test, X_test, combined_data)

        print(f'Accuracy: {accuracy}')
        total_acc += accuracy

    print('\n')
    print(f'Accumulated accuracy: {total_acc}')
    print(f'Total iterations: {count}')
    print(f'Average accuracy: {total_acc / count}')

# accuracy_run(3)
